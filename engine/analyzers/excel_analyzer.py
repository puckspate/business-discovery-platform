from pathlib import Path

from openpyxl import load_workbook


class ExcelAnalyzer:

    @staticmethod
    def analyze(file_path: str):

        workbook = load_workbook(
            filename=Path(file_path),
            read_only=True,
            data_only=True,
        )

        result = {
            "sheet_count": len(workbook.sheetnames),
            "sheet_names": workbook.sheetnames,
            "worksheets": []
        }

        total_rows = 0

        for sheet in workbook.sheetnames:

            ws = workbook[sheet]

            headers = []

            for cell in ws[1]:
                headers.append(cell.value)

            sheet_info = {
                "sheet_name": sheet,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "headers": headers
            }

            total_rows += ws.max_row

            result["worksheets"].append(sheet_info)

        result["total_rows"] = total_rows

        workbook.close()

        return result